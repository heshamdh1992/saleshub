from decimal import Decimal
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, Count, Avg, Q
from django.views.generic import TemplateView
from django.db.models import F
from apps.products.models import Product
from apps.sales.models import SaleItem
from apps.core.mixins import MerchantRequiredMixin
from apps.sales.models import Sale
from apps.customers.models import Customer
from apps.sales.models import Payment
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from apps.core.mixins import ManagerOrOwnerRequiredMixin

class SalesReportView(LoginRequiredMixin, ManagerOrOwnerRequiredMixin, TemplateView):
    template_name = "reports/sales_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        merchant = self.get_merchant()

        date_from = self.request.GET.get("date_from", "").strip()
        date_to = self.request.GET.get("date_to", "").strip()
        payment_type = self.request.GET.get("payment_type", "").strip()
        payment_status = self.request.GET.get("payment_status", "").strip()
        payment_currency = self.request.GET.get("payment_currency", "").strip()
        query = self.request.GET.get("q", "").strip()

        sales = Sale.objects.filter(
            merchant=merchant
        ).select_related(
            "customer", "payment_currency", "pricing_currency"
        ).order_by("-created_at")

        if date_from:
            sales = sales.filter(created_at__date__gte=date_from)

        if date_to:
            sales = sales.filter(created_at__date__lte=date_to)

        if payment_type:
            sales = sales.filter(payment_type=payment_type)

        if payment_status:
            sales = sales.filter(payment_status=payment_status)

        if payment_currency:
            sales = sales.filter(payment_currency__code=payment_currency)

        if query:
            sales = sales.filter(
                Q(invoice_number__icontains=query) |
                Q(customer__name__icontains=query)
            )

        summary = sales.aggregate(
            total_sales=Sum("total_amount"),
            total_profit=Sum("total_profit"),
            total_invoices=Count("id"),
            avg_invoice=Avg("total_amount"),
            total_due=Sum("amount_due"),
        )

        context.update({
            "sales": sales[:200],  # حد منطقي في البداية
            "total_sales": summary["total_sales"] or Decimal("0"),
            "total_profit": summary["total_profit"] or Decimal("0"),
            "total_invoices": summary["total_invoices"] or 0,
            "avg_invoice": summary["avg_invoice"] or Decimal("0"),
            "total_due": summary["total_due"] or Decimal("0"),

            "date_from": date_from,
            "date_to": date_to,
            "selected_payment_type": payment_type,
            "selected_payment_status": payment_status,
            "selected_payment_currency": payment_currency,
            "query": query,
        })

        return context
    
class SalesReportExportExcelView(LoginRequiredMixin, ManagerOrOwnerRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        merchant = self.get_merchant()

        date_from = request.GET.get("date_from", "").strip()
        date_to = request.GET.get("date_to", "").strip()
        payment_type = request.GET.get("payment_type", "").strip()
        payment_status = request.GET.get("payment_status", "").strip()
        payment_currency = request.GET.get("payment_currency", "").strip()
        query = request.GET.get("q", "").strip()

        sales = Sale.objects.filter(
            merchant=merchant
        ).select_related(
            "customer", "payment_currency", "pricing_currency"
        ).order_by("-created_at")

        if date_from:
            sales = sales.filter(created_at__date__gte=date_from)

        if date_to:
            sales = sales.filter(created_at__date__lte=date_to)

        if payment_type:
            sales = sales.filter(payment_type=payment_type)

        if payment_status:
            sales = sales.filter(payment_status=payment_status)

        if payment_currency:
            sales = sales.filter(payment_currency__code=payment_currency)

        if query:
            sales = sales.filter(
                Q(invoice_number__icontains=query) |
                Q(customer__name__icontains=query)
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        headers = [
            "Invoice Number",
            "Date",
            "Customer",
            "Payment Type",
            "Payment Status",
            "Pricing Currency",
            "Payment Currency",
            "Exchange Rate",
            "Subtotal",
            "Discount",
            "Total Amount USD",
            "Total Amount Payment Currency",
            "Amount Paid",
            "Amount Due",
            "Total Profit",
            "Notes",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for sale in sales:
            ws.append([
                sale.invoice_number,
                sale.created_at.strftime("%Y-%m-%d %H:%M"),
                sale.customer.name if sale.customer else "",
                sale.get_payment_type_display(),
                sale.get_payment_status_display(),
                sale.pricing_currency.code if sale.pricing_currency else "",
                sale.payment_currency.code if sale.payment_currency else "",
                float(sale.exchange_rate or 0),
                float(sale.subtotal or 0),
                float(sale.discount_amount or 0),
                float(sale.total_amount or 0),
                float(sale.total_amount_payment_currency or 0),
                float(sale.amount_paid or 0),
                float(sale.amount_due or 0),
                float(sale.total_profit or 0),
                sale.notes or "",
            ])

        column_widths = {
            "A": 18,
            "B": 20,
            "C": 22,
            "D": 15,
            "E": 18,
            "F": 16,
            "G": 16,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 18,
            "L": 24,
            "M": 14,
            "N": 14,
            "O": 14,
            "P": 30,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "sales_report.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response    
    
class ProductReportView(LoginRequiredMixin, ManagerOrOwnerRequiredMixin, TemplateView):
    template_name = "reports/product_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        merchant = self.get_merchant()

        date_from = self.request.GET.get("date_from", "").strip()
        date_to = self.request.GET.get("date_to", "").strip()
        sort_by = self.request.GET.get("sort_by", "qty").strip()

        products = Product.objects.filter(
            merchant=merchant,
            is_active=True
        ).order_by("name")

        stats_list = []

        for product in products:
            sale_items = SaleItem.objects.filter(
                sale__merchant=merchant,
                product=product
            )

            if date_from:
                sale_items = sale_items.filter(sale__created_at__date__gte=date_from)

            if date_to:
                sale_items = sale_items.filter(sale__created_at__date__lte=date_to)

            totals = sale_items.aggregate(
                total_qty=Sum("quantity"),
                total_sales=Sum("line_total"),
                total_cost=Sum(F("unit_cost") * F("quantity")),
            )

            total_qty = totals["total_qty"] or 0
            total_sales = totals["total_sales"] or Decimal("0")
            total_cost = totals["total_cost"] or Decimal("0")
            total_profit = total_sales - total_cost

            stock_quantity = product.stock_quantity or 0
            reorder_level = product.reorder_level or 0

            if stock_quantity <= 0:
                stock_status = "out"
            elif stock_quantity <= reorder_level:
                stock_status = "low"
            else:
                stock_status = "ok"

            stats_list.append({
                "product_id": product.id,
                "product_name": product.name,
                "stock_quantity": stock_quantity,
                "reorder_level": reorder_level,
                "stock_status": stock_status,
                "total_qty": total_qty,
                "total_sales": total_sales,
                "total_cost": total_cost,
                "total_profit": total_profit,
            })

        if sort_by == "profit":
            stats_list.sort(key=lambda x: x["total_profit"], reverse=True)
        elif sort_by == "sales":
            stats_list.sort(key=lambda x: x["total_sales"], reverse=True)
        elif sort_by == "stock":
            stats_list.sort(key=lambda x: x["stock_quantity"])
        else:
            stats_list.sort(key=lambda x: x["total_qty"], reverse=True)

        total_products = products.count()
        low_stock_count = products.filter(
            stock_quantity__gt=0,
            stock_quantity__lte=F("reorder_level")
        ).count()
        out_of_stock_count = products.filter(
            stock_quantity__lte=0
        ).count()

        context.update({
            "stats": stats_list,
            "total_products": total_products,
            "low_stock_count": low_stock_count,
            "out_of_stock_count": out_of_stock_count,
            "date_from": date_from,
            "date_to": date_to,
            "sort_by": sort_by,
        })

        return context

class CustomerReportView(LoginRequiredMixin, ManagerOrOwnerRequiredMixin, TemplateView):
    template_name = "reports/customer_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        merchant = self.get_merchant()
        query = self.request.GET.get("q", "").strip()
        sort_by = self.request.GET.get("sort_by", "balance").strip()

        customers = Customer.objects.filter(
            merchant=merchant,
            is_active=True
        )

        if query:
            customers = customers.filter(
                Q(name__icontains=query) |
                Q(phone__icontains=query)
            )

        customers = customers.prefetch_related("sales", "payments").order_by("name")

        rows = []
        total_sales_all = Decimal("0")
        total_payments_all = Decimal("0")
        total_balance_all = Decimal("0")
        indebted_customers_count = 0

        for customer in customers:
            total_sales = customer.total_invoices_amount()
            total_payments = customer.total_payments_amount()
            balance = customer.balance()

            sales_count = customer.sales.count()

            last_sale = customer.sales.order_by("-created_at").first()
            last_payment = customer.payments.order_by("-created_at").first()

            last_activity = None
            if last_sale and last_payment:
                last_activity = max(last_sale.created_at, last_payment.created_at)
            elif last_sale:
                last_activity = last_sale.created_at
            elif last_payment:
                last_activity = last_payment.created_at

            if balance > 0:
                indebted_customers_count += 1

            total_sales_all += total_sales
            total_payments_all += total_payments
            total_balance_all += balance

            rows.append({
                "customer": customer,
                "sales_count": sales_count,
                "total_sales": total_sales,
                "total_payments": total_payments,
                "balance": balance,
                "last_activity": last_activity,
            })

        if sort_by == "sales":
            rows.sort(key=lambda x: x["total_sales"], reverse=True)
        elif sort_by == "payments":
            rows.sort(key=lambda x: x["total_payments"], reverse=True)
        elif sort_by == "invoices":
            rows.sort(key=lambda x: x["sales_count"], reverse=True)
        elif sort_by == "activity":
            rows.sort(key=lambda x: x["last_activity"] or "", reverse=True)
        else:
            rows.sort(key=lambda x: x["balance"], reverse=True)

        context.update({
            "rows": rows,
            "query": query,
            "sort_by": sort_by,
            "total_customers": len(rows),
            "total_sales_all": total_sales_all,
            "total_payments_all": total_payments_all,
            "total_balance_all": total_balance_all,
            "indebted_customers_count": indebted_customers_count,
        })

        return context
    

class ProductReportExportExcelView(LoginRequiredMixin, ManagerOrOwnerRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        merchant = self.get_merchant()

        date_from = request.GET.get("date_from", "").strip()
        date_to = request.GET.get("date_to", "").strip()
        sort_by = request.GET.get("sort_by", "qty").strip()

        sale_items = SaleItem.objects.filter(
            sale__merchant=merchant
        ).select_related("product", "sale")

        if date_from:
            sale_items = sale_items.filter(sale__created_at__date__gte=date_from)

        if date_to:
            sale_items = sale_items.filter(sale__created_at__date__lte=date_to)

        product_stats = (
            sale_items.values(
                "product__id",
                "product__name",
                "product__sku",
                "product__barcode",
                "product__stock_quantity",
                "product__reorder_level",
            )
            .annotate(
                total_qty=Sum("quantity"),
                total_sales=Sum("line_total"),
                total_cost=Sum(F("unit_cost") * F("quantity")),
            )
        )

        stats_list = []
        for row in product_stats:
            total_sales = row["total_sales"] or Decimal("0")
            total_cost = row["total_cost"] or Decimal("0")
            total_profit = total_sales - total_cost

            stock_quantity = row["product__stock_quantity"] or 0
            reorder_level = row["product__reorder_level"] or 0

            if stock_quantity <= 0:
                stock_status = "Out of Stock"
            elif stock_quantity <= reorder_level:
                stock_status = "Low Stock"
            else:
                stock_status = "OK"

            stats_list.append({
                "product_name": row["product__name"],
                "sku": row["product__sku"] or "",
                "barcode": row["product__barcode"] or "",
                "stock_quantity": stock_quantity,
                "reorder_level": reorder_level,
                "stock_status": stock_status,
                "total_qty": row["total_qty"] or 0,
                "total_sales": total_sales,
                "total_cost": total_cost,
                "total_profit": total_profit,
            })

        if sort_by == "profit":
            stats_list.sort(key=lambda x: x["total_profit"], reverse=True)
        elif sort_by == "sales":
            stats_list.sort(key=lambda x: x["total_sales"], reverse=True)
        elif sort_by == "stock":
            stats_list.sort(key=lambda x: x["stock_quantity"])
        else:
            stats_list.sort(key=lambda x: x["total_qty"], reverse=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Product Report"

        headers = [
            "Product Name",
            "SKU",
            "Barcode",
            "Stock Quantity",
            "Reorder Level",
            "Stock Status",
            "Sold Quantity",
            "Total Sales USD",
            "Total Cost USD",
            "Total Profit USD",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for item in stats_list:
            ws.append([
                item["product_name"],
                item["sku"],
                item["barcode"],
                item["stock_quantity"],
                item["reorder_level"],
                item["stock_status"],
                item["total_qty"],
                float(item["total_sales"]),
                float(item["total_cost"]),
                float(item["total_profit"]),
            ])

        column_widths = {
            "A": 28,
            "B": 16,
            "C": 20,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 16,
            "H": 18,
            "I": 18,
            "J": 18,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "product_report.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response    
    

class CustomerReportExportExcelView(LoginRequiredMixin, ManagerOrOwnerRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        merchant = self.get_merchant()
        query = request.GET.get("q", "").strip()
        sort_by = request.GET.get("sort_by", "balance").strip()

        customers = Customer.objects.filter(
            merchant=merchant,
            is_active=True
        )

        if query:
            customers = customers.filter(
                Q(name__icontains=query) |
                Q(phone__icontains=query)
            )

        customers = customers.prefetch_related("sales", "payments").order_by("name")

        rows = []
        for customer in customers:
            total_sales = customer.total_invoices_amount()
            total_payments = customer.total_payments_amount()
            balance = customer.balance()

            sales_count = customer.sales.count()

            last_sale = customer.sales.order_by("-created_at").first()
            last_payment = customer.payments.order_by("-created_at").first()

            last_activity = None
            if last_sale and last_payment:
                last_activity = max(last_sale.created_at, last_payment.created_at)
            elif last_sale:
                last_activity = last_sale.created_at
            elif last_payment:
                last_activity = last_payment.created_at

            rows.append({
                "customer_name": customer.name,
                "phone": customer.phone or "",
                "sales_count": sales_count,
                "total_sales": total_sales,
                "total_payments": total_payments,
                "balance": balance,
                "last_activity": last_activity,
            })

        if sort_by == "sales":
            rows.sort(key=lambda x: x["total_sales"], reverse=True)
        elif sort_by == "payments":
            rows.sort(key=lambda x: x["total_payments"], reverse=True)
        elif sort_by == "invoices":
            rows.sort(key=lambda x: x["sales_count"], reverse=True)
        elif sort_by == "activity":
            rows.sort(key=lambda x: x["last_activity"] or "", reverse=True)
        else:
            rows.sort(key=lambda x: x["balance"], reverse=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Report"

        headers = [
            "Customer Name",
            "Phone",
            "Invoices Count",
            "Total Sales USD",
            "Total Payments USD",
            "Balance USD",
            "Last Activity",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([
                row["customer_name"],
                row["phone"],
                row["sales_count"],
                float(row["total_sales"]),
                float(row["total_payments"]),
                float(row["balance"]),
                row["last_activity"].strftime("%Y-%m-%d %H:%M") if row["last_activity"] else "",
            ])

        column_widths = {
            "A": 28,
            "B": 18,
            "C": 16,
            "D": 18,
            "E": 18,
            "F": 16,
            "G": 22,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "customer_report.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
